# Section1: Inserting and Getting Book data #
def CodeMaker(LanguagesAbbreviations, CategoryAbbreviations, Language, Category, Spreadsheet):
    #makes a unique ID for a book using its language, category and its spreadsheet.
        from openpyxl import load_workbook
        if Language in LanguagesAbbreviations and Category in CategoryAbbreviations:
            Current_Workbook = load_workbook(Spreadsheet)
            Current_Worksheet = Current_Workbook[Category]
            Last_Occupied_Row = Current_Worksheet.max_row
            Id_Number = Last_Occupied_Row + 1
            Language_Ab = LanguagesAbbreviations[Language]
            Category_Ab = CategoryAbbreviations[Category]
            Code = f'{Language_Ab}-{Category_Ab}/{Id_Number}'
            print(Code)
            return Code
        else:
            return 1

def InsertBook(BookName, Price, Author, Publisher, Audience, Language, Catagory):
    # inserts a book into a spreadsheet using its data (It takes Languages like "Malayalam" not M).
    Prohibited = ['', ' ']

    # TODO: Outsource the following three variables.
    Languages_Abbreviations = {'English': 'E', 'Malayalam': 'M', 'Hindi': 'H'}
    Spreadsheet_Addresses = {'M': 'Spreadsheets\Book_Details\Malayalam_books.xlsx', 'E': 'Spreadsheets\Book_Details\English_books.xlsx', 'H': 'Spreadsheets\Book_Details\Hindi_books.xlsx'}
    Category_Abbreviations = {'Fiction': 'FC', 'NonFiction': 'NF', 'SelfHelp': 'SF', 'Magazine': 'MG'}

    if BookName not in Prohibited and Author not in Prohibited and Publisher not in Prohibited:
        if Catagory.lower() != 'select' and Language != 'select':
            if Language in Languages_Abbreviations and Catagory in Category_Abbreviations:
                from openpyxl import load_workbook
                workbook = load_workbook(Spreadsheet_Addresses[Languages_Abbreviations[Language]])
                worksheet = workbook[Catagory]
                Book_Details = [BookName, Price,
                               CodeMaker(Languages_Abbreviations, Category_Abbreviations, Language,
                                         Catagory, Spreadsheet_Addresses[Languages_Abbreviations[Language]]), Author,
                               Publisher, Audience]
                if not Book_Details[2] == 1:
                    worksheet.append(Book_Details)
                    print(Book_Details)
                else:
                    return [1, "The language or the category is wrong :("]
            else:
                return 1
        else:
            return 1
    else:
        return 1

def FetchBookDetails(BookCode):
    #Returns a list of book detail from a book code.
    # can be used to autofill details into forms
    book_details = []
    from openpyxl import load_workbook
    Categories = {'FC': 'Fiction', 'NF': 'NonFiction', 'SH': 'SelfHelp', 'MG': 'Magazine'}
    Languages = {'M': 'Functionality\Malayalam_books.xlsx', 'E': 'Functionality/English_books.xlsx'}
    while True:
        if str(type(BookCode)) == "<class 'str'>":
            if not BookCode == '' and not BookCode == ' ':
                if '-' in BookCode and '/' in BookCode:
                    BookCode = BookCode.split('-')
                    LanguageIndex = BookCode[1].split('/')
                    BookCode[1] = LanguageIndex[0]
                    BookCode.append(LanguageIndex[1])
                    BookCode[1] = BookCode[1].upper()
                    BookCode[2] = BookCode[2].upper()
                    if BookCode[0] in Languages:
                        if BookCode[1] in Categories:
                            workbook = load_workbook(Languages[BookCode[0]])
                            worksheet = workbook[Categories[BookCode[1]]]
                            for cell in range(1, 7):
                                book_details.append(worksheet.cell(int(BookCode[2]) + 1, column=cell).value)
                            return book_details
                        else:
                            return [1, 'The catagory is wrong :(']
                    else:
                        return [1, 'The language code is not specified:(']
                else:
                    return [1, 'The code is not correctly written :(']
            else:
                return [1, "You Didn't write any code :("]
        else:
            return [1, 'Please only enter a string']
# --End of Section 1-- #

# Section2: Book Statistics
def FindOverdueBooks():
    from openpyxl import load_workbook
    import datetime
    borrowed_book_details = []
    Borrowed_Books_Data = load_workbook("Borrowed_books.xlsx")
    DataSheet = Borrowed_Books_Data['DataSheet']
    for row in range(2, DataSheet.max_row + 1):
        returned = DataSheet.cell(row=row, column=6).value
        if not returned:
            R_Date = DataSheet.cell(row=row, column=3).value
            R_Month = DataSheet.cell(row=row, column=4).value
            R_Year = DataSheet.cell(row=row, column=5).value
            Original_Return_Date = datetime.date(R_Year, R_Month, R_Date)
            today = datetime.date.today()
            print(today)
            time_left_for_return = Original_Return_Date - today
            if time_left_for_return.days <= 0:
                Book_Details = []
                for column in range(1, 3):
                    Book_Details.append(DataSheet.cell(row=row, column=column).value)
                borrowed_book_details.append(Book_Details)
    print(borrowed_book_details)
# --End of section 2--