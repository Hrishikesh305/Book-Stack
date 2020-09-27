# Section1: Member Statistics and Processed Details

#--End Of Section 1--
#Section2: Book Borrowing
def BorrowBookbyMember(BorrowerID, ListOfBookCodes):
    ProhibitedCharecters = ['', ' ']
    ColumnsNotNeeded = [1, 4] # Columns in the member details spreadsheet
    BorrowerDetails = []
    BookAndErrorDetails = []
    try:
        int(BorrowerID)
    except ValueError:
        return [1, 'Only type in a number :(']
    BorrowerID = int(BorrowerID)
    from openpyxl import load_workbook
    AllMembersDetails = load_workbook("Spreadsheets\Members_deails.xlsx")
    Sheet = AllMembersDetails['Sheet1']
    for column in range(1, 7):
        if not column in ColumnsNotNeeded:
            cell = Sheet.cell(BorrowerID + 1, column)
            if not cell.value == None: 
                BorrowerDetails.append(cell.value)
            else:
                return [1, 'The Member Does not exist :(']
        else:
            pass
    #Prepare the borrowed book Details    
    import Functionality.BMT
    ListOfBookCodes = ListOfBookCodes.split(', ')
    for BookCode in ListOfBookCodes:
        BookDetails = Functionality.BMT.FetchBookDetails(str(BookCode))
        BookAndErrorDetails.append(BookDetails)

    # Empties textfile from last codes
    f = open('BasicProgramData\TemoraryStorage\BorrowerDetails.txt', 'r+')
    f.truncate(0)
    f.close()
    # Append details    
    with open("BasicProgramData\TemoraryStorage\BorrowerDetails.txt", 'w') as TempFile:
        for Detail in BorrowerDetails:
            TempFile.write(str(Detail) + ',')
        BookNumber = 0
        for Book in BookAndErrorDetails:
            TempFile.write(f'\n')
            if not Book[0] == 1:
                for Detail in Book:
                    TempFile.write(str(Detail) + ',')
            else:
                return BookAndErrorDetails
                
    return [0, 'no problem']