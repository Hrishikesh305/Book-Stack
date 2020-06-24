def CodeMaker(Languages_Abbreviations, Category_Abbreviations, Language, Category, Spreadsheet):
        from openpyxl import load_workbook
        if Language in Languages_Abbreviations and Category in Category_Abbreviations:
            Current_Workbook = load_workbook(Spreadsheet)
            Current_Worksheet = Current_Workbook[Category]
            Last_Occupied_Row = Current_Worksheet.max_row
            Id_Number = Last_Occupied_Row + 1
            Language_Ab = Languages_Abbreviations[Language]
            Category_Ab = Category_Abbreviations[Category]
            Code = f'{Language_Ab}-{Category_Ab}/{Id_Number}'
            return Code
        else:
            return 1
        

def insert_book(Book_Name, Price, Author, Publisher, Made_For, Language, Catagory):

    Prohibited = ['', ' ']

    # TODO: Outsource the following three variables.
    Languages_Abbreviations = {'English': 'E', 'Malayalam': 'M', 'Hindi': 'H'}
    Spreadsheet_Addresses = {'M': 'Malayalam_books.xlsx', 'E': 'English_books.xlsx', 'H': 'Hindi_books.xlsx'}
    Category_Abbreviations = {'Fiction': 'FC', 'NonFiction': 'NF', 'SelfHelp': 'SF', 'Magazine': 'MG'}

    if Book_Name not in Prohibited and Author not in Prohibited and Publisher not in Prohibited:
        if Catagory.lower() != 'select' and Language != 'select':
            if Language in Languages_Abbreviations and Catagory in Category_Abbreviations:
                from openpyxl import load_workbook
                workbook = load_workbook(Spreadsheet_Addresses[Languages_Abbreviations[Language]])
                worksheet = workbook[Catagory]
                Book_Details = [Book_Name, Price,
                               CodeMaker(Languages_Abbreviations, Category_Abbreviations, Language,
                                         Catagory, Spreadsheet_Addresses[Languages_Abbreviations[Language]]), Author,
                               Publisher, Made_For]
                if not Book_Details[2] == 1:
                    worksheet.append(Book_Details)
                    print(Book_Details)
                else:
                    return 1
            else:
                return 1
        else:
            return 1
    else:
        return 1

def FetchBookDetails(Book_Code):
#Returns a list of book detail from a book code.
    book_details = []
    from openpyxl import load_workbook
    Categories = {'FC': 'Fiction', 'NF': 'NonFiction', 'SH': 'SelfHelp', 'MG': 'Magazine'}
    Languages = {'M': 'Malayalam_books.xlsx', 'E': 'English_books.xlsx'}
    while True:
        if str(type(Book_Code)) == "<class 'str'>":
            if not Book_Code == '' and not Book_Code == ' ':
                if '-' in Book_Code and '/' in Book_Code:
                    Book_Code = Book_Code.split('-')
                    LanguageIndex = Book_Code[1].split('/')
                    Book_Code[1] = LanguageIndex[0]
                    Book_Code.append(LanguageIndex[1])
                    Book_Code[1] = Book_Code[1].upper()
                    Book_Code[2] = Book_Code[2].upper()
                    if Book_Code[0] in Languages:
                        if Book_Code[1] in Categories:
                            workbook = load_workbook(Languages[Book_Code[0]])
                            worksheet = workbook[Categories[Book_Code[1]]]
                            for cell in range(1, 7):
                                book_details.append(worksheet.cell(int(Book_Code[2]) + 1, column=cell).value)
                            return book_details
                        else:
                            return [1, 'The catagory is wrong :(']
                    else:
                        return [1, 'The language code is not specified:(']
                else:
                    return [1, 'The code is not correctly written :(']
            else:
                return [1, "You Didn't write any code :("]
        else:
            return [1, 'Please only enter a string']
