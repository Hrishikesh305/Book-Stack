# TODO: Optimize this code
from openpyxl import load_workbook
from random import choice

workbook = load_workbook('Functionality/Malayalam_books.xlsx')
Worksheet = workbook['Sheet1']
all_values = []
all_positions = []
Other_Popular_Books = []
Popular_Books = []
for row in range(2, Worksheet.max_row + 1):
    cell = Worksheet.cell(row, 3)
    all_values.append(cell.value)
    all_positions.append(row)
# tier one
# print(all_values)
biggest_value = max(all_values)
# print(f'biggest_value: {biggest_value}')
FirstPopularBookPosition = all_values.index(biggest_value) + 2
FirstPopularBookName = Worksheet.cell(FirstPopularBookPosition, 1).value
# print(f'Book name: {FirstPopularBookName}')
for row in range(2, Worksheet.max_row + 1):
    if Worksheet.cell(row=row, column=3).value == biggest_value:
        Other_Popular_Books.append(Worksheet.cell(row, 1).value)

if Other_Popular_Books.count(Other_Popular_Books) <= 3:
    for counter in range(1, 4):
        Popular_Books.append(choice(Other_Popular_Books))
print(Popular_Books)

